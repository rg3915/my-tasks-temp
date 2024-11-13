import json
import os
import re
import subprocess
from collections import defaultdict
from datetime import date, datetime, timedelta

import gitlab
import requests
from decouple import config
from django.utils import timezone
from openpyxl import Workbook, load_workbook, styles
from rich import print
from rich.console import Console

from backend.core.utils import datetime_to_string, timedelta_to_string
from backend.task.models import Issue, Label, Sprint, Task, Timesheet

console = Console()


FOLDER_BASE = '/home/regis/Dropbox/projetos'


CONJUGATIONS = {
    'adicionar': 'adiciona',
    'aplicar': 'aplica',
    'cadastrar': 'cadastra',
    'calcular': 'calcula',
    'corrigir': 'corrige',
    'criar': 'cria',
    'deletar': 'deleta',
    'editar': 'edita',
    'entregar': 'entrega',
    'enviar': 'envia',
    'gerar': 'gera',
    'imprimir': 'imprime',
    'instalar': 'instala',
    'processar': 'processa',
    'remover': 'remove',
    'retornar': 'retorna',
    'salvar': 'salva',
}


def conjugate_infinitive(sentence):
    # Função para substituir verbos no infinitivo pela forma conjugada
    def replace_verb(match):
        verb = match.group(0)
        conjugated = CONJUGATIONS.get(verb.lower(), verb)  # Busca a forma conjugada em minúsculas

        # Mantém a capitalização se a primeira letra do verbo original for maiúscula
        if verb[0].isupper():
            conjugated = conjugated.capitalize()

        return conjugated

    # Padrão para encontrar verbos no infinitivo
    infinitive_pattern = r"\b(\w+ar|\w+er|\w+ir)\b"

    # Substituir verbos no infinitivo pela conjugação correta
    conjugated_sentence = re.sub(infinitive_pattern, replace_verb, sentence)

    return conjugated_sentence


def check_if_the_date_already_exists(filename, milestone_title):
    date_format = datetime.now().strftime('%Y-%m-%d')

    # Check if the file exists, and create it if it doesn't
    if not os.path.exists(filename):
        with open(filename, 'w') as file:
            file.write(f"## {milestone_title}\n")

    # Check if the date already exists in the file
    with open(filename, 'r') as file:
        file_contents = file.read()
        if date_format in file_contents:
            print("Date already exists in the file.")
        else:
            # Append the date to the file
            with open(filename, 'a') as file:
                file.write(f"\n## {date_format}\n\n")
            print(f"Date {date_format} added to the file {filename}.")


def write_tarefas(task):
    # Escreve lorem na segunda linha depois do match.
    # n;n;n; significa três next.
    sprint = Sprint.objects.filter(project=task.project).last()
    tarefas_filename = f'{FOLDER_BASE}/{sprint.project.customer.name}/{sprint.project.title}/tarefas.txt'

    task = f'{task.issue.number} - {task.title}'

    lorem = 'lorem'
    sed_command = 'sed -i "/\\] ' + task + '/{n;n;n;s/.*/' + lorem + '/}" ' + tarefas_filename
    subprocess.run(sed_command, shell=True)

    # Substitui lorem por AQUI
    asterisco = '\\n    AQUI'
    sed_command = f'sed -i "s/{lorem}/{asterisco}\\n/" {tarefas_filename}'
    subprocess.run(sed_command, shell=True)


def remove_aqui_from_tarefas(task):
    """
    Remove AQUI de tarefas.txt
    """
    sprint = Sprint.objects.filter(project=task.project).last()
    tarefas_filename = f'{FOLDER_BASE}/{sprint.project.customer.name}/{sprint.project.title}/tarefas.txt'

    # sed_command = f'sed -i "s/    AQUI/{{N;d;}}" {tarefas_filename}'
    sed_command = f'sed -i "/    AQUI/{{N;d;}}" {tarefas_filename} && sed -i "s/    AQUI//" {tarefas_filename}'
    subprocess.run(sed_command, shell=True)


def write_x_on_tarefas(task):
    sprint = Sprint.objects.filter(project=task.project).last()
    tarefas_filename = f'{FOLDER_BASE}/{sprint.project.customer.name}/{sprint.project.title}/tarefas.txt'

    task_text = f'{task.issue.number} - {task}'
    # print(task_text)
    escaped_task_text = task_text.replace("/", "\\/")  # Escape any slashes in the task_text
    command = f"sed -i 's/\\[ \\] {escaped_task_text}/\\[x\\] {escaped_task_text}/' {tarefas_filename}"
    subprocess.run(command, shell=True, check=True)


def write_on_tarefas(filename, issue, labels, is_bug):
    today = date.today().strftime('%d/%m/%y')

    _labels = ','.join(labels)

    with open(filename, 'a') as f:
        f.write('\n---\n\n')
        f.write(f'[ ] {issue.number} - {issue.title}\n')
        f.write(f'    {_labels}\n')
        f.write(f'    {today}\n\n\n')

        if issue.description:
            f.write(f'    {issue.description}\n\n')

        title = conjugate_infinitive(issue.title)
        if is_bug:
            title = f'bugfix: {title}'

        project = issue.sprint.project.title
        customer = issue.sprint.project.customer.name
        milestone = issue.milestone.title

        if customer == 'euroled':
            f.write('    python ~/gitlab/my-tasks/backend/core/write_changelog_euroled.py\n')
            f.write(f'    echo "* {title}. #{issue.number}" >> ~/{customer}/CHANGELOG.md\n')
        elif customer == 'ledsoft':
            f.write(f'    echo "* {title}. #{issue.number}" >> ~/my/{project}/CHANGELOG.md\n')
        else:
            f.write(f'    echo "* {title}. #{issue.number}" >> ~/{customer}/{project}/CHANGELOG.md\n')

        f.write(f'    echo "* {title}. #{issue.number}" >> ~/Dropbox/projetos/{customer}/{project}/changelog/CHANGELOG_{milestone}.md\n\n')  # noqa E501

        f.write(f"    cd ~/gitlab/my-tasks; sa; m start_task -p='{project}' -t={issue.number} -ph=True\n")

        if project == 'ekoospregao':
            f.write(
                f"    workon {customer}; python cli/task.py -c start -p {project} -t {issue.number} --previous_hour\n")

        f.write(f"\n    _gadd '{title}. close #{issue.number}'; # gp\n\n")

        f.write(f"    cd ~/gitlab/my-tasks; sa; m stop_task -p='{project}' -t={issue.number}\n")

        if project == 'ekoospregao':
            f.write(f"    workon {customer}; python cli/task.py -c end -p {project} -t {issue.number}\n")


def write_changelog_dropbox(issue):
    customer = issue.sprint.project.customer.name
    project = issue.sprint.project.title
    filename = f'{FOLDER_BASE}/{customer}/{project}/changelog/CHANGELOG_{issue.milestone.title}.md'
    print(filename)

    check_if_the_date_already_exists(filename, issue.milestone.title)

    text = f'* {issue.title}. #{issue.number}\n'

    with open(filename, 'r') as f:
        content = f.read()
        # Escreve somente se o texto não existir.
        if text not in content:
            with open(filename, 'a') as f:
                f.write(text)


def save_issue(data):
    labels = Label.objects.filter(label__in=data['labels'])
    sprint = Sprint.objects.filter(project=data['project']).last()

    issue = Issue.objects.create(
        number=data['iid'],
        title=data['title'],
        description=data['description'],
        milestone=data['milestone'],
        sprint=sprint,
        url=data['web_url'],
    )
    for label in labels:
        issue.labels.add(label)

    filename = f'{FOLDER_BASE}/{sprint.project.customer.name}/{sprint.project.title}/tarefas.txt'

    number = data['iid']
    title = data['title']

    print(filename)
    print(number)
    print(title)

    is_bug = False
    if 'bug' in data['labels']:
        is_bug = True

    write_on_tarefas(filename, issue, data['labels'], is_bug)
    return issue


def save_issue_multiple(datas):
    issues = []

    sorted_data = sorted(datas, key=lambda x: x['iid'])

    for data in sorted_data:
        labels = Label.objects.filter(label__in=data['labels'])
        sprint = Sprint.objects.filter(project=data['project']).last()

        issue, _ = Issue.objects.get_or_create(
            number=data['iid'],
            title=data['title'],
            description=data['description'],
            milestone=data['milestone'],
            sprint=sprint,
            url=data['web_url'],
        )

        issues.append(issue)

        for label in labels:
            issue.labels.add(label)

        filename = f'{FOLDER_BASE}/{sprint.project.customer.name}/{sprint.project.title}/tarefas.txt'

        number = data['iid']
        title = data['title']

        print(filename)
        print(number)
        print(title)

        is_bug = False
        if 'bug' in data['labels']:
            is_bug = True

        write_on_tarefas(filename, issue, data['labels'], is_bug)
    return issues


def update_issue(data):
    labels = Label.objects.filter(label__in=data['labels'])
    sprint = Sprint.objects.filter(project=data['project']).last()

    issue = Issue.objects.filter(number=data['iid'], sprint__project=data['project']).first()

    payload = dict(
        number=data['iid'],
        title=data['title'],
        description=data['description'],
        milestone=data['milestone'],
    )

    for attr, value in payload.items():
        setattr(issue, attr, value)

    issue.save()

    print(issue)

    issue.labels.clear()
    for label in labels:
        issue.labels.add(label)

    filename = f'{FOLDER_BASE}/{sprint.project.customer.name}/{sprint.project.title}/tarefas.txt'

    is_bug = False
    if 'bug' in data['labels']:
        is_bug = True

    write_on_tarefas(filename, issue, data['labels'], is_bug)
    return issue


def save_task(issue):
    Task.objects.get_or_create(
        title=issue.title,
        project=issue.sprint.project,
        issue=issue,
    )


def save_task_multiple(issues):
    for issue in issues:
        save_task(issue)


def update_task(issue):
    task = Task.objects.filter(issue=issue).first()
    task.title = issue.title
    task.save()


def create_timesheet(task, previous_hour):
    now = datetime.now()
    start_time = datetime.now()

    if previous_hour:
        # TODO: verificar se é do mesmo projeto.
        last_hour = Timesheet.objects.first()  # é o último por causa da ordenação.
        start_time = last_hour.end_time
        print('start:', datetime_to_string(start_time - timedelta(hours=3), '%H:%M'))
        print('now:', datetime_to_string(now, '%H:%M'))
    else:
        print(datetime_to_string(start_time, '%H:%M'))

    return Timesheet.objects.create(task=task, start_time=start_time)


def stop_timesheet(task):
    now = datetime.now()

    # print(datetime_to_string(now, '%H:%M'))

    # Na verdade não precisa da task.
    timesheet = Timesheet.objects.filter(task=task, end_time__isnull=True).first()
    timesheet.end_time = now
    timesheet.save()

    end_time = timezone.make_aware(timesheet.end_time)  # Convert to aware datetime

    print()
    total = str(end_time - timesheet.start_time)
    print(f'Total: {total.split(".")[0]}')
    print()

    return timesheet


def create_gitlab_issue(args):
    '''
    Requer /etc/rg3915.cfg
    '''
    gl = gitlab.Gitlab.from_config('somewhere', ['/etc/rg3915.cfg'])

    title, body, labels, project, milestone = args.values()

    gl_project = gl.projects.get(project.gitlab_project_id)

    data_dict = {
        "title": f"{title}",
        "description": f"{body}",
        "assignee_id": config('GITLAB_ASSIGNEE_ID'),
        "labels": labels,
        "milestone_id": milestone.original_id,
    }

    response = gl_project.issues.create(data_dict)

    data = json.loads(response.to_json())

    # TESTE
    # data = {
    #     "iid": 7,
    #     "title": "Teste",
    #     "description": "Descrição teste.",
    #     "labels": ["backend", "frontend"],
    #     # "time_stats":
    #     # {
    #     #     "time_estimate": 0,
    #     #     "total_time_spent": 0,
    #     # },
    #     "web_url": "https://gitlab.com/rg3915/my-tasks/-/issues/7",
    # }

    data['project'] = project
    data['milestone'] = milestone

    return data


def create_github_issue(args):
    title, body, labels, project, milestone = args.values()

    assignee = 'rg3915'

    repo_owner = project.repository_owner
    repo_name = project.title
    token = project.github_token
    URL = f'https://api.github.com/repos/{repo_owner}/{repo_name}/issues'

    headers = {"Authorization": f"token {token}"}

    labels = labels.split(',')

    # Create our issue
    issue = {
        "title": title,
        "body": body,
        "assignees": [assignee],
        "labels": labels,
        "milestone": milestone.original_id,
    }

    # Add the issue to our repository
    response = requests.post(URL, headers=headers, json=issue)

    if response.status_code == 201:
        console.print(f'Successfully created Issue "{title}"', style='green')
        data = response.json()

        data['project'] = project
        data['labels'] = labels
        data['iid'] = data['number']
        data['description'] = data['body']
        data['milestone'] = milestone
        data['web_url'] = data['html_url']

        return data

    console.print(f'Could not create Issue "{title}"', style='red')


def update_gitlab_issue(args):
    '''
    Requer /etc/rg3915.cfg
    '''
    gl = gitlab.Gitlab.from_config('somewhere', ['/etc/rg3915.cfg'])

    id, title, body, labels, project, milestone = args.values()

    gl_project = gl.projects.get(project.gitlab_project_id)

    data_dict = {
        "title": f"{title}",
        "description": f"{body}",
        "assignee_id": config('GITLAB_ASSIGNEE_ID'),
        "labels": labels,
        "milestone_id": milestone.original_id,
    }

    response = gl_project.issues.update(id, data_dict)

    data = response

    data['project'] = project
    data['milestone'] = milestone

    return data


def update_github_issue(args):
    issue, title, body, labels, project, milestone = args.values()

    assignee = 'rg3915'

    repo_owner = project.repository_owner
    repo_name = project.title
    token = project.github_token
    URL = f'https://api.github.com/repos/{repo_owner}/{repo_name}/issues/{issue}'

    headers = {"Authorization": f"token {token}"}

    labels = labels.split(',')

    issue = {
        "title": title,
        "body": body,
        "assignees": [assignee],
        "labels": labels,
        "milestone": milestone.original_id,
    }

    # Add the issue to our repository
    response = requests.post(URL, headers=headers, json=issue)

    if response.status_code == 200:
        console.print(f'Successfully updated Issue "{title}"', style='green')
        data = response.json()

        data['project'] = project
        data['labels'] = labels
        data['iid'] = data['number']
        data['description'] = data['body']
        data['milestone'] = milestone
        data['web_url'] = data['html_url']

        return data

    console.print(f'Could not update Issue "{title}"', style='red')


def read_gitlab_issue(args):
    '''
    Requer /etc/rg3915.cfg
    '''
    gl = gitlab.Gitlab.from_config('somewhere', ['/etc/rg3915.cfg'])

    milestone, assignee, project = args.values()

    gl_project = gl.projects.get(project.gitlab_project_id)

    response = gl_project.issues.list(per_page=100, order_by='created_at')
    # response = gl_project.issues.list(per_page=100, milestone='0.0.1', state='opened')
    # response = gl_project.issues.list(per_page=100, order_by='created_at')

    items = []
    for data in response:
        _data = {}
        _data['iid'] = data.iid
        _data['title'] = data.title
        _data['description'] = data.description
        _data['labels'] = data.labels
        _data['web_url'] = data.web_url
        _data['project'] = project
        _data['milestone'] = milestone
        items.append(_data)

    return items


def read_github_issue(args):
    milestone, assignee, project = args.values()

    repo_owner = project.repository_owner
    repo_name = project.title
    token = project.github_token
    URL = f'https://api.github.com/repos/{repo_owner}/{repo_name}/issues'

    headers = {"Authorization": f"token {token}"}

    # Read issues
    response = requests.get(URL, headers=headers)

    issues = []
    if response.status_code == 200:
        console.print('Successfully read Issues', style='green')

        for data in response.json():
            labels = [item['name'] for item in data['labels']]
            data['project'] = project
            data['labels'] = labels
            data['iid'] = data['number']
            data['description'] = data['body']
            data['milestone'] = milestone
            data['web_url'] = data['html_url']

            issues.append(data)

        return issues

    console.print(f'Could not read Issues', style='red')


def get_hour_display(_time):
    hours, remainder = divmod(_time.total_seconds(), 3600)
    minutes = divmod(remainder, 60)
    time_str = ''

    if hours:
        time_str += f'{int(hours)}h '

    if minutes[0]:
        time_str += f'{int(minutes[0])}m'

    if not time_str:
        return '0'

    return time_str.strip()


def group_by_date(project):
    '''
    Agrupa os dados por dia.
    '''
    timesheet_data = Timesheet.objects.filter(task__project__title=project).values(
        'start_time__date',
        'task__issue__number',
        'start_time',
        'end_time',
        'task__issue__sprint__number',
    ).order_by('start_time')

    # Cria um dicionário para armazenar as horas totais e as issues por data.
    result_dict = defaultdict(lambda: {'total_hours': timedelta(), 'issues': set()})

    for timesheet in timesheet_data:
        start_time = timesheet.get('start_time')
        end_time = timesheet.get('end_time')

        total_hours = timedelta()
        if start_time is not None and end_time is not None:
            total_hours = end_time - start_time

        date_only = timesheet['start_time'].date()
        issues = timesheet['task__issue__number']
        result_dict[date_only]['total_hours'] += total_hours
        result_dict[date_only]['issues'].add(issues)
        result_dict[date_only]['sprint'] = timesheet['task__issue__sprint__number']

    output = [
        {
            'date': datetime_to_string(key, '%d/%m/%y'),
            'month': key.month,
            'total_hours': value['total_hours'],
            'total_hours_display': str(get_hour_display(value['total_hours'])),
            'issues': ', '.join(map(str, value['issues'])),
            'sprint': value['sprint'],
        }
        for key, value in result_dict.items()
    ]

    return output


def group_data(project, group_by_field, key_field):
    timesheet_data = Timesheet.objects.filter(task__project__title=project).values(
        'start_time',
        'end_time',
        group_by_field
    ).order_by('start_time')

    # Cria um dicionário para armazenar as horas totais e as issues por data.
    result_dict = defaultdict(lambda: {'total_hours': timedelta(), 'issues': set()})

    for timesheet in timesheet_data:
        # Agrupa pelo campo definido em group_by_field.
        date_only = timesheet[group_by_field]

        start_time = timesheet.get('start_time')
        end_time = timesheet.get('end_time')

        total_hours = timedelta()
        if start_time is not None and end_time is not None:
            total_hours = end_time - start_time

        result_dict[date_only]['total_hours'] += total_hours

    output = [
        {
            key_field: key,
            'total_hours': value['total_hours'],
            'total_hours_display': str(get_hour_display(value['total_hours'])),
        }
        for key, value in result_dict.items()
    ]

    return output


def group_by_month(project):
    '''
    Agrupa os dados por mês.
    '''
    return group_data(project, 'start_time__month', 'month')


def group_by_sprint(project):
    '''
    Agrupa os dados por sprint.
    '''
    return group_data(project, 'task__issue__sprint__number', 'sprint')


def write_total_hours_on_timesheet_file(timesheet_filename, total_hours):
    wb = load_workbook(timesheet_filename)
    try:
        ws = wb['total_hours']
    except KeyError:
        ws = wb.create_sheet('total_hours')

    new_row = 2

    labels = (
        'data',
        'mês',
        'tempo',
        'tempo_display',
        'issues',
        'sprint',
    )

    bold_calibri = styles.Font(bold=True, name='Calibri')

    # Set labels and apply font in a loop
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold_calibri

    for item in total_hours:
        ws.cell(row=new_row, column=1, value=item['date'])

        cell = ws.cell(row=new_row, column=2, value=item['month'])
        cell.alignment = styles.Alignment(horizontal='center')

        ws.cell(row=new_row, column=3, value=item['total_hours'])
        ws.cell(row=new_row, column=4, value=item['total_hours_display'])
        ws.cell(row=new_row, column=5, value=item['issues'])

        cell = ws.cell(row=new_row, column=6, value=item['sprint'])
        cell.alignment = styles.Alignment(horizontal='center')

        new_row += 1

    wb.save(timesheet_filename)


def write_total_hours(timesheet_filename, total_hours, group_by_field, start, columns):
    wb = load_workbook(timesheet_filename)
    try:
        ws = wb['total_hours']
    except KeyError:
        ws = wb.create_sheet('total_hours')

    new_row = 2

    labels = (
        group_by_field,
        'tempo',
        'tempo_display',
    )

    bold_calibri = styles.Font(bold=True, name='Calibri')

    # Set labels and apply font in a loop
    for col, label in enumerate(labels, start=start):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold_calibri

    for item in total_hours:
        cell = ws.cell(row=new_row, column=columns[0], value=item[group_by_field])
        cell.alignment = styles.Alignment(horizontal='center')

        ws.cell(row=new_row, column=columns[1], value=item['total_hours'])
        ws.cell(row=new_row, column=columns[2], value=item['total_hours_display'])

        new_row += 1

    wb.save(timesheet_filename)


def write_total_hours_by_month_on_timesheet_file(timesheet_filename, total_hours):
    return write_total_hours(
        timesheet_filename=timesheet_filename,
        total_hours=total_hours,
        group_by_field='month',
        start=8,
        columns=(8, 9, 10)
    )


def write_total_hours_by_sprint_on_timesheet_file(timesheet_filename, total_hours):
    return write_total_hours(
        timesheet_filename=timesheet_filename,
        total_hours=total_hours,
        group_by_field='sprint',
        start=12,
        columns=(12, 13, 14)
    )


def export_timesheet_service(project):
    customer = project.customer.name
    title = project.title
    timesheet_filename = f'{FOLDER_BASE}/{customer}/{project}/timesheet_teste_{title}.xlsx'

    print(f'Salvo em {timesheet_filename}')

    try:
        wb = load_workbook(timesheet_filename)
        ws = wb['timesheet']
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.create_sheet('timesheet')
        wb.save(timesheet_filename)

    new_row = 2

    labels = (
        'data',
        'hora_inicial',
        'hora_final',
        'tempo',
        'tempo_display',
        'issue',
        'titulo',
    )

    bold_calibri = styles.Font(bold=True, name='Calibri')

    # Set labels and apply font in a loop
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold_calibri

    for timesheet in project.get_tasks():
        ws.cell(row=new_row, column=1, value=timesheet.date_from_start_time_display)
        ws.cell(row=new_row, column=2, value=timesheet.start_time_display_fixed_hour)
        ws.cell(row=new_row, column=3, value=timesheet.end_time_display_fixed_hour)
        ws.cell(row=new_row, column=4, value=timesheet.get_hour())
        ws.cell(row=new_row, column=5, value=timesheet.get_hour_display())

        cell = ws.cell(row=new_row, column=6, value=timesheet.task.issue.number)
        cell.alignment = styles.Alignment(horizontal='center')

        ws.cell(row=new_row, column=7, value=timesheet.task.issue.title)

        new_row += 1

    wb.save(timesheet_filename)

    total_hours = group_by_date(project)

    write_total_hours_on_timesheet_file(timesheet_filename, total_hours)

    total_hours = group_by_month(project)

    write_total_hours_by_month_on_timesheet_file(timesheet_filename, total_hours)

    total_hours = group_by_sprint(project)

    write_total_hours_by_sprint_on_timesheet_file(timesheet_filename, total_hours)
