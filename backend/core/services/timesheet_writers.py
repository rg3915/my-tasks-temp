from openpyxl import load_workbook, styles


def write_total_hours_on_timesheet_file(timesheet_filename, total_hours):
    wb = load_workbook(timesheet_filename)
    try:
        ws = wb['total_hours']
    except KeyError:
        ws = wb.create_sheet('total_hours')

    new_row = 2

    labels = (
        'data',
        'mês',
        'tempo',
        'tempo_display',
        'issues',
        'sprint',
    )

    bold_calibri = styles.Font(bold=True, name='Calibri')

    # Set labels and apply font in a loop
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold_calibri

    for item in total_hours:
        ws.cell(row=new_row, column=1, value=item['date'])

        cell = ws.cell(row=new_row, column=2, value=item['month'])
        cell.alignment = styles.Alignment(horizontal='center')

        ws.cell(row=new_row, column=3, value=item['total_hours'])
        ws.cell(row=new_row, column=4, value=item['total_hours_display'])
        ws.cell(row=new_row, column=5, value=item['issues'])

        cell = ws.cell(row=new_row, column=6, value=item['sprint'])
        cell.alignment = styles.Alignment(horizontal='center')

        new_row += 1

    wb.save(timesheet_filename)


def write_total_hours(timesheet_filename, total_hours, group_by_field, start, columns):
    wb = load_workbook(timesheet_filename)
    try:
        ws = wb['total_hours']
    except KeyError:
        ws = wb.create_sheet('total_hours')

    new_row = 2

    labels = (
        group_by_field,
        'tempo',
        'tempo_display',
    )

    bold_calibri = styles.Font(bold=True, name='Calibri')

    # Set labels and apply font in a loop
    for col, label in enumerate(labels, start=start):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold_calibri

    for item in total_hours:
        cell = ws.cell(row=new_row, column=columns[0], value=item[group_by_field])
        cell.alignment = styles.Alignment(horizontal='center')

        ws.cell(row=new_row, column=columns[1], value=item['total_hours'])
        ws.cell(row=new_row, column=columns[2], value=item['total_hours_display'])

        new_row += 1

    wb.save(timesheet_filename)


def write_total_hours_by_month_on_timesheet_file(timesheet_filename, total_hours):
    return write_total_hours(
        timesheet_filename=timesheet_filename,
        total_hours=total_hours,
        group_by_field='month',
        start=8,
        columns=(8, 9, 10),
    )


def write_total_hours_by_sprint_on_timesheet_file(timesheet_filename, total_hours):
    return write_total_hours(
        timesheet_filename=timesheet_filename,
        total_hours=total_hours,
        group_by_field='sprint',
        start=12,
        columns=(12, 13, 14),
    )
