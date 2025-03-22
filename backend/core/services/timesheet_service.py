from datetime import datetime, timedelta

from django.utils import timezone
from openpyxl import Workbook, load_workbook, styles
from rich import print

from backend.core.services.data_groupers import group_by_date, group_by_month, group_by_sprint
from backend.core.services.timesheet_writers import (
    write_total_hours_by_month_on_timesheet_file,
    write_total_hours_by_sprint_on_timesheet_file,
    write_total_hours_on_timesheet_file,
)
from backend.core.utils import datetime_to_string
from backend.task.models import Timesheet

FOLDER_BASE = '/home/regis/Dropbox/projetos'


def create_timesheet(task):
    # Obtém o momento atual com informação de timezone
    now = timezone.now()

    # Busca o último registro de timesheet (assumindo ordenação por mais recente)
    last_hour = Timesheet.objects.first()

    if last_hour:
        # Calcula a diferença de tempo entre agora e o último registro
        # Ambos os horários precisam ter informação de timezone para fazer a subtração
        time_diff = now - last_hour.end_time

        # Converte a diferença de tempo para minutos para facilitar a comparação
        diff_minutes = time_diff.total_seconds() / 60

        # Se a diferença for menor que 15:01 (15 minutos e 1 segundo),
        # usa o horário de término do último registro como início

        # usa 15.02 para considerar possíveis milissegundos
        # Verifica se é o mesmo projeto.
        if diff_minutes < 15.02 and task.project == last_hour.task.project:
            start_time = last_hour.end_time
        else:
            # Se passou muito tempo, inicia com o horário atual
            start_time = now

        print('start:', datetime_to_string(start_time - timedelta(hours=3), '%H:%M'))
        print('now:', datetime_to_string(now - timedelta(hours=3), '%H:%M'))
    else:
        # Se não existe registro anterior, usa o horário atual
        start_time = now
        print(datetime_to_string(start_time, '%H:%M'))

    # Cria e retorna um novo registro de timesheet
    return Timesheet.objects.create(task=task, start_time=start_time)


def stop_timesheet(task):
    now = datetime.now()

    # print(datetime_to_string(now, '%H:%M'))

    # Na verdade não precisa da task.
    timesheet = Timesheet.objects.filter(task=task, end_time__isnull=True).first()
    timesheet.end_time = now
    timesheet.save()

    end_time = timezone.make_aware(timesheet.end_time)  # Convert to aware datetime

    print()
    total = str(end_time - timesheet.start_time)
    print(f'Total: {total.split(".")[0]}')
    print()

    return timesheet


def export_timesheet_service(project):
    customer = project.customer.name
    title = project.title
    timesheet_filename = f'{FOLDER_BASE}/{customer}/{project}/timesheet_teste_{title}.xlsx'

    print(f'Salvo em {timesheet_filename}')

    try:
        wb = load_workbook(timesheet_filename)
        ws = wb['timesheet']
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.create_sheet('timesheet')
        wb.save(timesheet_filename)

    new_row = 2

    labels = (
        'data',
        'hora_inicial',
        'hora_final',
        'tempo',
        'tempo_display',
        'issue',
        'titulo',
    )

    bold_calibri = styles.Font(bold=True, name='Calibri')

    # Set labels and apply font in a loop
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold_calibri

    for timesheet in project.get_tasks():
        ws.cell(row=new_row, column=1, value=timesheet.date_from_start_time_display)
        ws.cell(row=new_row, column=2, value=timesheet.start_time_display_fixed_hour)
        ws.cell(row=new_row, column=3, value=timesheet.end_time_display_fixed_hour)
        ws.cell(row=new_row, column=4, value=timesheet.get_hour())
        ws.cell(row=new_row, column=5, value=timesheet.get_hour_display())

        cell = ws.cell(row=new_row, column=6, value=timesheet.task.issue.number)
        cell.alignment = styles.Alignment(horizontal='center')

        ws.cell(row=new_row, column=7, value=timesheet.task.issue.title)

        new_row += 1

    wb.save(timesheet_filename)

    total_hours = group_by_date(project)

    write_total_hours_on_timesheet_file(timesheet_filename, total_hours)

    total_hours = group_by_month(project)

    write_total_hours_by_month_on_timesheet_file(timesheet_filename, total_hours)

    total_hours = group_by_sprint(project)

    write_total_hours_by_sprint_on_timesheet_file(timesheet_filename, total_hours)
